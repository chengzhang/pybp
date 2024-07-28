"""sample to read write xlsx file"""


from typing import List, Dict, Text
import openpyxl
import sys


def dump(sentences, file_path):
    """to label format, store local"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value='sensitive label')
    sheet.cell(row=1, column=2, value='sentence')
    excel_row = 2
    for sentence in sentences:
        try:
            sheet.cell(row=excel_row, column=2, value=sentence)
            excel_row += 1
        except:
            continue
    workbook.save(file_path)


def load(in_file):
    """load from in_file"""
    workbook = openpyxl.load_workbook(in_file)
    sheet = workbook.active
    for i, row in enumerate(sheet.rows):
        for j, cell in enumerate(row):
            print(i, j, cell.value, end=" ")


def dump_xlsx_table(table, filename, header=None):
    """ table: itr of row """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row_idx = 1
    if header:
        for j, value in header:
            sheet.cell(row=row_idx, column=j+1, value=value)
        row_idx += 1
    for row in table:
        for j, value in row:
            sheet.cell(row=row_idx, column=j+1, value=value)
        row_idx += 1
    workbook.save(filename)


def dump_multi_sheet(name_to_sentences: Dict[Text, List[Text]], filepath: Text):
    wb = openpyxl.Workbook()
    for name, sentences in name_to_sentences.items():
        sheet = wb.create_sheet(name)
        for row_idx, sentence in enumerate(sentences):
            sheet.cell(row=row_idx+1, column=1, value=sentence)
    wb.save(filepath)
